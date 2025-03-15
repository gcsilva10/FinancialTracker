# tracker/api_views.py
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from rest_framework.generics import RetrieveDestroyAPIView
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework.parsers import MultiPartParser, FormParser
from .serializers import TransactionSerializer
from rest_framework.generics import RetrieveAPIView
from .models import Transaction


class TransactionCreateAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        serializer = TransactionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(user=request.user)  # associa a transação ao usuário autenticado
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TransactionDetailAPIView(RetrieveDestroyAPIView):
    serializer_class = TransactionSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Retorna apenas as transações do usuário autenticado
        return Transaction.objects.filter(user=self.request.user)
    
class TransactionsByMonthAPIView(APIView):
#Retorna todas as transações do usuário autenticado para o mês/ano especificado. O parâmetro de query esperado é 'month' no formato 'YYYY-MM'. Se não for fornecido, retorna todas as transações.
    
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        month_param = request.query_params.get('month', None)
        if month_param:
            try:
                year_str, month_str = month_param.split('-')
                year = int(year_str)
                month = int(month_str)
            except Exception as e:
                return Response({'error': 'Formato inválido. Use YYYY-MM.'}, status=status.HTTP_400_BAD_REQUEST)
            transactions = Transaction.objects.filter(user=request.user, date__year=year, date__month=month)
        else:
            transactions = Transaction.objects.filter(user=request.user)

        serializer = TransactionSerializer(transactions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class ExportTransactionsExcelAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        # Recupera as transações do usuário autenticado
        transactions = Transaction.objects.filter(user=request.user)
        
        # Cria uma nova planilha Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transações"

        # Cria o cabeçalho
        headers = ['ID', 'Tipo', 'Valor', 'Categoria', 'Descrição', 'Data']
        ws.append(headers)

        # Adiciona os dados
        for t in transactions:
            row = [
                t.id,
                t.type,
                t.amount,
                t.category,
                t.description,
                t.date.strftime('%Y-%m-%d')
            ]
            ws.append(row)

        # Salva a planilha em um buffer de memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Retorna a resposta com o arquivo Excel
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transacoes.xlsx"'
        return response
    

class ImportTransactionsExcelAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]  # Permite upload de arquivo

    def post(self, request, format=None):
        file_obj = request.FILES.get('file', None)
        if not file_obj:
            return Response({"error": "Nenhum arquivo enviado."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            # Carrega a planilha
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active

            # Supondo que a primeira linha seja o cabeçalho
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            # Cabeçalhos esperados: ['ID', 'Tipo', 'Valor', 'Categoria', 'Descrição', 'Data']
            # Vamos ignorar o ID e criar novas transações
            created_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Extrai os valores; assumindo a ordem: ID, type, amount, category, description, date
                type_val = row[1]
                amount_val = row[2]
                category_val = row[3]
                description_val = row[4] if row[4] is not None else ""
                date_val = row[5]  # Espera-se que esteja no formato "YYYY-MM-DD" ou um objeto date

                # Cria a transação para o usuário autenticado
                Transaction.objects.create(
                    user=request.user,
                    type=type_val,
                    amount=amount_val,
                    category=category_val,
                    description=description_val,
                    date=date_val
                )
                created_count += 1

            return Response({"message": f"Importadas {created_count} transações com sucesso."}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
